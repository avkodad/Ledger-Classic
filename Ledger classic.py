import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill

#Att fixa:
# cell witdh
#TYPO misstag


class Ledger:

    @classmethod
    def __init__(self):

        self.wb = load_workbook('Budget.xlsx')
        self.ws = self.wb.active
        self.mxcl = self.ws.max_column

 
        self.setup(self)



    def setup(self):
        while True:
            m = input('Vilken månad vill du budgetera?\n').title()

            for col in range(1,self.mxcl+1):
                
                månad = self.ws.cell(row=1, column = col).value

                
                if m == månad:

                    
                    lön = int(input('Vad får du ut denna månad?\n').replace(' ', ''))

                    l = self.ws.cell(row=2, column=col, value=lön)
                    l.font = Font(name='Arial', size=12, bold=True, color='6AEA4D')
                    self.wb.save('Budget.xlsx')

                    print ('{}kr i lön för {} månad'.format(lön, m))
                    self.ledger(lön,col)
                


    @classmethod
    def ledger(self, lön, col):
        nyLön = lön
        while True:
            
            for cell in range(3, 1000, 2):

                if (self.ws.cell(row=cell, column=col).value) is None:        


                    utgifter = {}
                    utgift = input('\nNamn på utgift?\n')


                    if utgift == 'x': #om klar med att fylla i uppgifter

                        
                        u = self.ws.cell(row=cell-1, column=col)
                        u.font = Font(name='Arial', size=12, bold=True, color='6AEA4D')

                        self.wb.save('Budget.xlsx')
                        print ('Tack för du organiserar din ekonomi med Ledger Classic')
                        exit()

                    
                    else:


                        pris = int(input('Pris i kronor?\n').replace(' ', ''))
                        #if type(pris) != int: 'bara siffror???'
                        utgifter[utgift]=pris
                        nyLön -= pris

                        print ('\n -{}kr:  {}\n{}kr'.format(pris, utgift, nyLön))
                        self.xldata(utgifter, nyLön, col, cell)


                else:


                    print ('Redan upptagen: {}'.format(self.ws.cell(row=cell, column=col).value))
                    
                    return



    @classmethod
    def xldata(self, utgifter   , nyLön, col, cell):
        for utgift, pris in utgifter.items():


            a = self.ws.cell(row=cell, column=col, value='-{}'.format(pris))
            a.font = Font(name='Arial', size=10, color='B22810')
            self.ws.cell(row=cell, column=col+1, value=utgift)#cell width?
            self.ws.cell(row=cell+1, column=col, value=nyLön)
            

        self.wb.save('Budget.xlsx')





Ledger()


##10695	
##-957	csn återkrav
##9738	
##-586	csn årsbelopp
##9152	
##-6000	hyra
##3152	
##-499	bjj
##2653	
##-350	klippning
##2303	
##-145	fastpris
##2158	
##-1000	mat
##1158	
